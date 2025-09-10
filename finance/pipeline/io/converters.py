import csv
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, List, Tuple, Optional

from openpyxl import load_workbook
import chardet


def xlsx_to_csv(
    input_xlsx: Path,
    output_csv: Path,
    sheet_name: Optional[str] = None,
    sheet_index: int = 0,
    auto_select_best: bool = True,
) -> None:
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(filename=str(input_xlsx), read_only=True, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif auto_select_best:
        # Heuristic: prefer a tab whose name looks like sales/royalty/detail
        # otherwise choose the sheet with most rows/cols
        def name_score(name: str) -> int:
            n = name.lower()
            score = 0
            for kw in ("sales", "royalty", "report", "detail"):
                if kw in n:
                    score += 10
            if "overview" in n or "summary" in n:
                score -= 5
            return score

        candidates = []
        for ws0 in wb.worksheets:
            try:
                r = ws0.max_row or 0
                c = ws0.max_column or 0
                candidates.append((name_score(ws0.title), r, c, ws0))
            except Exception:
                continue
        candidates.sort(key=lambda x: (x[0], x[1], x[2]), reverse=True)
        ws = candidates[0][3] if candidates else wb.worksheets[sheet_index]
    else:
        ws = wb.worksheets[sheet_index]
    with output_csv.open("w", encoding="utf-8", newline="") as f_out:
        writer = csv.writer(f_out, delimiter=",")
        for row in ws.iter_rows(values_only=True):
            writer.writerow([_cell_to_str(v) for v in row])
    wb.close()


def detect_encoding(input_file: Path, default: str = "utf-8") -> str:
    with input_file.open("rb") as f:
        raw = f.read(200000)
    guess = chardet.detect(raw)
    enc = (guess or {}).get("encoding") or default
    return enc


def normalize_delimiter_and_decimal(
    input_csv: Path,
    output_csv: Path,
    delimiter_in: str = ";",
    delimiter_out: str = ",",
    decimal_comma_to_dot: bool = True,
    encoding_in: Optional[str] = None,
) -> None:
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    if not encoding_in:
        encoding_in = detect_encoding(input_csv)
    with input_csv.open("r", encoding=encoding_in, errors="ignore", newline="") as f_in, output_csv.open(
        "w", encoding="utf-8", newline=""
    ) as f_out:
        reader = csv.reader(f_in, delimiter=delimiter_in)
        writer = csv.writer(f_out, delimiter=delimiter_out)
        for row in reader:
            if decimal_comma_to_dot:
                row = [cell.replace("\u00A0", " ").replace(",", ".") if _looks_like_number(cell) else cell for cell in row]
            writer.writerow(row)


def _looks_like_number(text: str) -> bool:
    if not text:
        return False
    t = text.strip()
    # quick heuristic to avoid touching dates and IDs with hyphens/letters
    return any(ch.isdigit() for ch in t) and all(ch.isdigit() or ch in ",.-" for ch in t)


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        try:
            return value.isoformat()
        except Exception:
            return str(value)
    return str(value)


